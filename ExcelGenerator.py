# Author
import openpyxl
import os


class ExcelGenerator:
    def __init__(self, file_name):
        if not os.path.exists(file_name):
            self.workbook = openpyxl.Workbook()
            try:
                self.workbook.save(file_name)
            except FileNotFoundError as exp:
                container_path = os.path.normpath(
                    file_name + os.path.sep + os.path.pardir)
                os.mkdir(container_path)
                self.workbook.save(file_name)
        else:
            self.workbook = openpyxl.load_workbook(file_name)
        self.file_name = file_name
        self.sheet = self.workbook.active

    def query_cell_value(self, cell_row, cell_column):
        return self.sheet.cell(row=cell_row, column=cell_column).value

    def insert_cell_value(self, cell_row, cell_column, cell_value):
        self.sheet.cell(row=cell_row, column=cell_column).value = cell_value
        
        return self.sheet.cell(row=cell_row, column=cell_column).value

    def insert_cell_hyperlink(self, cell_row, cell_column, cell_link):
        self.sheet.cell(row=cell_row, column=cell_column).hyperlink = cell_link
        
        return self.sheet.cell(row=cell_row, column=cell_column).hyperlink

    def save_workbook(self):
        self.workbook.save(self.file_name)
        
    def insert_row(self, row) :
        self.sheet.insert_row(row)

if __name__ == '__main__':
    excel_file = ExcelGenerator("test_excel/01_test_file.xlsx")
    # print(type(excel_file.insert_cell_value(1, 1, 12)))
    print(excel_file.query_cell_value(1, 1))
    print(excel_file.insert_cell_hyperlink(1, 1, os.path.normpath("C:/Users\ASUS\Desktop/2021-Projects\Python-Assett/02-Libraries-Implementation\excel").replace(os.path.sep, '/')))
